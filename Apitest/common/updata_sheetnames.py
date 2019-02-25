#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/12/29 15:37
# @Author  : updata_sheetnames
import xlrd
from openpyxl import load_workbook
#修改sheets名字方法
def updata_sheetnames(sheetsa = 'Sheet1'):
    filename = '../case/demo-text.xlsx'#获取路径
    wb = load_workbook(filename)#打开文件
    sheets1 = wb.sheetnames#获取sheet名称
    sheet1 = wb[sheets1[0]]#获取第一个sheet名称
    ws = sheet1
    ws.title = sheetsa  # 修改名为Sheet1工作表名称
    wb.save(filename)  # 保存变更
    wb.close()
if __name__=='__main__':
    updata_sheetnames()
