# -*- coding:utf-8 -*-
from openpyxl import load_workbook
import openpyxl
from Apitest.common import updata_sheetnames
#设置两个参数，文件路径
def copy_excel(excelpath1, excelpath2):
    '''复制excek，把excelpath1数据复制到excelpath2'''
    #打开文件
    wb2 = openpyxl.Workbook()
    #保存文件
    wb2.save(excelpath2)
    # 读取两个文件的数据
    wb1 = openpyxl.load_workbook(excelpath1)
    wb2 = openpyxl.load_workbook(excelpath2)
    #读取wb1的所有sheel
    sheets1 = wb1.sheetnames
    # 读取wb2的所有sheel
    sheets2 = wb2.sheetnames
    #读取wb1第一个sheel
    sheet1 = wb1[sheets1[0]]
    #读取wb2第一个sheel
    sheet2 = wb2[sheets2[0]]
    #获取第一个sheel里最大的行数
    max_row = sheet1.max_row         # 最大行数
    #获取第一个sheel里最大的列数
    max_column = sheet1.max_column   # 最大列数
    #遍历从1开始，到最大行数
    for m in list(range(1,max_row+1)):

        for n in list(range(97,97+max_column)):   # chr(97)='a'
            n = chr(n)                            # ASCII字符
            i ='%s%d'% (n, m)                     # 单元格编号
            cell1 = sheet1[i].value               # 获取data单元格数据
            sheet2[i].value = cell1               # 赋值到test单元格
    wb2.save(excelpath2)                 # 保存数据
    wb1.close()                          # 关闭excel
    wb2.close()
class Write_excel(object):
    '''修改excel数据'''
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active  # 激活sheet

    def write(self, row_n, col_n, value):
        '''写入数据，如(2,3，"hello"),第二行第三列写入数据"hello"'''
        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.filename)



if __name__ == "__main__":
    copy_excel('../data/textcase01.xlsx','../case/demo-text.xlsx')
    wt = Write_excel('../case/demo-text.xlsx')
    updata_sheetnames.updata_sheetnames()
    print('ok')